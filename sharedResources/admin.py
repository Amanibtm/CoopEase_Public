import traceback

import pandas as pd
from django.contrib import admin
from django.shortcuts import redirect, render
from django.urls import path
from django import forms
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from EducationalEstablishmentResources.scheduler import generate_schedule_for_departments
from .models import typeEtablissement, TargetedPublic, Department, Equipement, Day, TimeSlot, Space, typeSalle, Grade, Faculty


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField()


def get_filtered_queryset(request, model):
    """Returns a filtered queryset based on admin's scope"""
    if request.user.is_superuser:
        return model.objects.all()

    if request.user.role == 'est_admin':
        est = request.user.estAdminProfile.est
        if model == Department:
            return model.objects.filter(est=est).distinct()
        elif model == Faculty:
            return model.objects.filter(est=est).distinct()
        elif model == TargetedPublic:
            return model.objects.filter(id=est.id).distinct()
        elif model == Space:
            return model.objects.filter(departements__est=est).distinct()
        elif model == TimeSlot:
            return model.objects.filter(departements__est=est).distinct()
        elif model == Equipement:
            return model.objects.filter(departement__est=est).distinct()
        return model.objects.all()

    if request.user.role == 'dep_admin':
        dep = request.user.depAdminProfile.departement
        if model == Department:
            return model.objects.filter(id=dep.id).distinct()
        elif model == Faculty:
            return model.objects.filter(est=dep.est).distinct()
        elif model == TargetedPublic:
            return model.objects.filter(id=dep.est.id).distinct()
        elif model == Space:
            return model.objects.filter(departements=dep).distinct()
        elif model == TimeSlot:
            return model.objects.filter(departements=dep).distinct()
        elif model == Equipement:
            return model.objects.filter(departement=dep).distinct()
        return model.objects.all()

    if request.user.is_staff and request.user.role == 'advancedUser':
        dep = request.user.Advanceduser_profile.Role.departement
        if model == Department:
            return model.objects.filter(id=dep.id).distinct()
        elif model == Faculty:
            return model.objects.filter(est=dep.est).distinct()
        elif model == TargetedPublic:
            return model.objects.filter(id=dep.est.id).distinct()
        elif model == Space:
            return model.objects.filter(departements=dep).distinct()
        elif model == TimeSlot:
            return model.objects.filter(departements=dep).distinct()
        elif model == Equipement:
            return model.objects.filter(departement=dep).distinct()
        return model.objects.all()

    return model.objects.none()


@admin.register(typeEtablissement)
class typeEtablissementAdmin(admin.ModelAdmin):
    list_display = ('name',)
    list_filter = ('name',)
    search_fields = ('name',)

    def get_queryset(self, request):
        return get_filtered_queryset(request, typeEtablissement)


@admin.register(TargetedPublic)
class TargetedPublicAdmin(admin.ModelAdmin):
    list_display = ('name', 'type', 'address', 'SystemeEtudeEnGroupe')
    list_filter = ('name', 'type', 'address', 'SystemeEtudeEnGroupe')
    search_fields = ('name', 'type', 'address', 'SystemeEtudeEnGroupe')

    def save_model(self, request, obj, form, change):
        # If semester system is not enabled, clear the dates
        if not obj.systemSemestre:
            obj.debutsemestre = None
            obj.finsemestre = None
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        return get_filtered_queryset(request, TargetedPublic)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'type' in form.base_fields:
                form.base_fields['type'].queryset = get_filtered_queryset(request, typeEtablissement)
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Nom", "Type", "Adresse", "Système d'étude en groupe"]
            ws.append(headers)

            types = typeEtablissement.objects.all().values_list('name', flat=True)
            system_choices = ['Groupe-Groupe', 'Groupes-Tous']

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('B', types)  # Type
            add_dropdown('D', system_choices)  # Système d'étude en groupe

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Etablissements.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Nom': 'name',
                        'Type': 'type',
                        'Adresse': 'address',
                        "Système d'étude en groupe": "SystemeEtudeEnGroupe",
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            type_obj = typeEtablissement.objects.get(name=row['type']) if row['type'] else None
                            system_value = row['SystemeEtudeEnGroupe']

                            if not TargetedPublic.objects.filter(
                                    name=row['name'],
                                    type=type_obj,
                                    address=row['address'],
                                    SystemeEtudeEnGroupe=system_value
                            ).exists():
                                TargetedPublic.objects.create(
                                    name=row['name'],
                                    type=type_obj,
                                    address=row['address'],
                                    SystemeEtudeEnGroupe=system_value,
                                )
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })
        except Exception as e:
            self.message_user(request, f"Erreur: {str(e)}", level='error')
            return redirect("..")


@admin.register(Faculty)
class FacultyAdmin(admin.ModelAdmin):
    list_display = ('name','est')
    list_filter = ('name','est')
    search_fields = ('name','est')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Faculty)

@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ('name', 'est', 'groupsCapacity')
    list_filter = ('name', 'est')
    search_fields = ('name', 'est__name')
    actions = ['generate_schedule_s1', 'generate_schedule_s2', 'generate_schedule_no_semester']

    def get_queryset(self, request):
        return get_filtered_queryset(request, Department)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'est' in form.base_fields:
                form.base_fields['est'].queryset = get_filtered_queryset(request, TargetedPublic)
                if request.user.role == 'est_admin':
                    form.base_fields['est'].initial = request.user.estAdminProfile.est
                elif request.user.role == 'dep_admin':
                    form.base_fields['est'].initial = request.user.depAdminProfile.departement.est
                elif request.user.role == 'advancedUser':
                    form.base_fields['est'].initial = request.user.Advanceduser_profile.Role.departement.est
        return form

    def generate_schedule_s1(self, request, queryset):
        generate_schedule_for_departments(list(queryset), semester='s1')
        self.message_user(request, f"Planning générée pour S1 dans {queryset.count()} département(s).")

    generate_schedule_s1.short_description = "Générer planning S1 (Universités)"

    def generate_schedule_s2(self, request, queryset):
        generate_schedule_for_departments(list(queryset), semester='s2')
        self.message_user(request, f"Planning générée pour S2 dans {queryset.count()} département(s).")

    generate_schedule_s2.short_description = "Générer planning S2 (Universités)"

    def generate_schedule_no_semester(self, request, queryset):
        generate_schedule_for_departments(list(queryset), semester=None)
        self.message_user(request, f"Planning générée pour {queryset.count()} département(s) (Lycées/Écoles).")

    generate_schedule_no_semester.short_description = "Générer planning (Lycées/Écoles)"

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Établissement","Faculté", "Nom du département", "Taille de groupe optimale"]
            ws.append(headers)

            establishments = get_filtered_queryset(request, TargetedPublic).values_list('name', flat=True)
            facultés = get_filtered_queryset(request, Faculty).values_list('name', flat=True)

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', establishments)
            add_dropdown('B', facultés)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Departements.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Établissement': 'est',
                        'Faculté': 'faculty',
                        'Nom du département': 'name',
                        'Taille de groupe optimale': 'groupsCapacity',
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            est_obj = TargetedPublic.objects.get(name=row['est']) if row['est'] else None
                            fac_obj = Faculty.objects.get(name=row['faculty']) if row['faculty'] else None
                            if not Department.objects.filter(est=est_obj,faculty=fac_obj,name=row['name'],
                                                            groupsCapacity=row['groupsCapacity']).exists():
                                Department.objects.create(est=est_obj,faculty=fac_obj,name=row['name'],groupsCapacity=row['groupsCapacity'],)
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            print(f"Unexpected error in import_excel: {traceback.format_exc()}")
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(typeSalle)
class typeSalleAdmin(admin.ModelAdmin):
    list_display = ('name',)
    list_filter = ('name',)
    search_fields = ('name',)

    def get_queryset(self, request):
        return get_filtered_queryset(request, typeSalle)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'name':
            kwargs[
                'help_text'] = "Un type de salle similaire existe déjà."
        return super().formfield_for_dbfield(db_field, request, **kwargs)


@admin.register(Space)
class SpaceAdmin(admin.ModelAdmin):
    list_display = ('room_number', 'room_type', 'capacity', 'get_departements')
    list_filter = ('room_number', 'room_type')
    search_fields = ('room_number', 'room_type__name', 'capacity', 'departements__name')

    def get_departements(self, obj):
        return ", ".join([d.name for d in obj.departements.all()])

    get_departements.short_description = 'Départements'

    def get_queryset(self, request):
        return get_filtered_queryset(request, Space)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'room_type' in form.base_fields:
                form.base_fields['room_type'].queryset = get_filtered_queryset(request, typeSalle)
            if 'departements' in form.base_fields:
                faculties = get_filtered_queryset(request, Faculty)
                fac_depts = Department.objects.filter(faculty__in=faculties)
                form.base_fields['departements'].queryset = fac_depts if fac_depts.exists() else get_filtered_queryset(request, Department)
                if request.user.role == 'dep_admin':
                    form.base_fields['departements'].initial = request.user.depAdminProfile.departement
                elif request.user.role == 'advancedUser':
                    form.base_fields['departements'].initial = request.user.Advanceduser_profile.Role.departement
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Réf", "Capacité", "Type de salle", "Départements"]
            ws.append(headers)

            room_types = typeSalle.objects.all().values_list('name', flat=True)
            faculties = get_filtered_queryset(request, Faculty)
            fac_depts = Department.objects.filter(faculty__in=faculties).distinct()
            if fac_depts.exists():
                departments = fac_depts.values_list('name', flat=True)
            else:
                departments = get_filtered_queryset(request, Department).values_list('name', flat=True)

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('C', room_types)  # Type de salle
            add_dropdown('D', departments)  # Départements

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Salles.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Réf': 'room_number',
                        'Capacité': 'capacity',
                        "Type de salle": "room_type",
                        "Départements": "departements",
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            room_type_obj = typeSalle.objects.get(name=row['room_type']) if row['room_type'] else None
                            departement_names = [name.strip() for name in row['departements'].split(',')]
                            departement_objs = Department.objects.filter(name__in=departement_names)

                            if not Space.objects.filter(
                                    room_number=row['room_number'],
                                    room_type=room_type_obj,
                                    capacity=row['capacity']
                            ).exists():
                                space = Space.objects.create(
                                    room_number=row['room_number'],
                                    room_type=room_type_obj,
                                    capacity=row['capacity'],
                                )
                                space.departements.set(departement_objs)
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            print(f"Unexpected error in import_excel: {traceback.format_exc()}")
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(Equipement)
class EquipementAdmin(admin.ModelAdmin):
    list_display = ('type', 'Brand', 'Nserie', 'departement', 'Status')
    list_filter = ('type', 'Brand', 'Nserie', 'departement', 'Status')
    search_fields = ('type', 'Brand', 'Nserie', 'departement__name')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Equipement)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'departement' in form.base_fields:
                form.base_fields['departement'].queryset = get_filtered_queryset(request, Department)
                if request.user.role == 'dep_admin':
                    form.base_fields['departement'].initial = request.user.depAdminProfile.departement
                elif request.user.role == 'advancedUser':
                    form.base_fields['departement'].initial = request.user.Advanceduser_profile.Role.departement
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Département","Type d'equipement", "Numéro de série", "Matricule", "Marque", "État"]
            ws.append(headers)

            departments = get_filtered_queryset(request, Department).values_list('name', flat=True)
            status_choices = ['Fonctionnel', 'En panne', 'En réparation', 'Hors service']

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', departments)  # Département
            add_dropdown('F', status_choices)  # État

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Equipements.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Département': 'departement',
                        "Type d'equipement": 'type',
                        'Numéro de série': 'Nserie',
                        'Matricule': 'Matricule',
                        "Marque": "Brand",
                        "État": "Status",
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            departement_obj = Department.objects.get(name=row['departement']) if row['departement'] else None
                            if not Equipement.objects.filter(
                                    departement=departement_obj,
                                    type=row['type'],
                                    Nserie=row['Nserie'],
                                    Matricule=row['Matricule'],
                                    Brand=row['Brand'],
                                    Status=row['Status']
                            ).exists():
                                Equipement.objects.create(
                                    departement=departement_obj,
                                    type=row['type'],
                                    Nserie=row['Nserie'],
                                    Matricule=row['Matricule'],
                                    Brand=row['Brand'],
                                    Status=row['Status'],
                                )
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            print(f"Unexpected error in import_excel: {traceback.format_exc()}")
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(Day)
class DayAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

    def get_queryset(self, request):
        return get_filtered_queryset(request, Day)


@admin.register(TimeSlot)
class TimeSlotAdmin(admin.ModelAdmin):
    list_display = ('start_time', 'end_time', 'get_departements', 'get_days')
    list_filter = ('start_time', 'end_time')
    search_fields = ('start_time', 'end_time')

    def get_departements(self, obj):
        return ", ".join([d.name for d in obj.departements.all()])

    get_departements.short_description = 'Départements'

    def get_days(self, obj):
        return ", ".join([d.name for d in obj.day.all()])

    get_days.short_description = 'Jours'

    def get_queryset(self, request):
        return get_filtered_queryset(request, TimeSlot)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'departements' in form.base_fields:
                form.base_fields['departements'].queryset = get_filtered_queryset(request, Department)
                if request.user.role == 'dep_admin':
                    form.base_fields['departements'].initial = request.user.depAdminProfile.departement
                elif request.user.role == 'advancedUser':
                    form.base_fields['departements'].initial = request.user.Advanceduser_profile.Role.departement
            if 'day' in form.base_fields:
                form.base_fields['day'].queryset = get_filtered_queryset(request, Day)
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Départements", "Jours", "Heure de début", "Heure de fin"]
            ws.append(headers)

            departments = get_filtered_queryset(request, Department).values_list('name', flat=True)
            days = Day.objects.all().values_list('name', flat=True)

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', departments)  # Départements
            add_dropdown('B', days)  # Jours

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=CreneauxHoraires.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Départements': 'departements',
                        'Jours': 'day',
                        'Heure de début': 'start_time',
                        "Heure de fin": "end_time",
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            departement_names = [name.strip() for name in row['departements'].split(',')]
                            departement_objs = Department.objects.filter(name__in=departement_names)
                            day_names = [name.strip() for name in row['day'].split(',')]
                            day_objs = Day.objects.filter(name__in=day_names)

                            if not TimeSlot.objects.filter(
                                    start_time=row['start_time'],
                                    end_time=row['end_time']
                            ).exists():
                                timeslot = TimeSlot.objects.create(
                                    start_time=row['start_time'],
                                    end_time=row['end_time'],
                                )
                                timeslot.departements.set(departement_objs)
                                timeslot.day.set(day_objs)
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
                print(f"Unexpected error in import_excel: {traceback.format_exc()}")
                self.message_user(request, "An unexpected error occurred", level='error')
                return redirect("..")

@admin.register(Grade)
class GradeAdmin(admin.ModelAdmin):
    list_display = ('name', 'teachCours', 'teachTD', 'teachTP')
    list_filter = ('name', 'teachCours', 'teachTD', 'teachTP')
    search_fields = ('name', 'teachCours', 'teachTD', 'teachTP')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Grade)