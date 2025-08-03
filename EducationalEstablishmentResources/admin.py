from django.db.models.functions import Substr
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.template.loader import get_template
from django.urls import path
from django import forms
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from sharedResources.models import Department, typeSalle
from users.models import AdvancedUserProfile
from .create_assign_groups import groupCreation, assignToGroups
from .models import Level, Speciality, Group, Module, Schedule
from EducationalEstablishmentResources.schedulerSpecialities import generate_schedule_for_specialities
import pandas as pd

from django.contrib import admin


admin.site.site_header = "Horix Admin"       # Title in the browser tab



class ExcelImportForm(forms.Form):
    excel_file = forms.FileField()


def get_filtered_queryset(request, model):
    """Returns a filtered queryset based on admin's scope"""
    if request.user.is_superuser:
        return model.objects.all()

    if request.user.role == 'est_admin':
        est = request.user.estAdminProfile.est
        if model == Level:
            return model.objects.filter(departement__est=est).distinct()
        elif model == Speciality:
            return model.objects.filter(level__departement__est=est).distinct()
        elif model == Group:
            return model.objects.filter(speciality__level__departement__est=est).distinct()
        elif model == Module:
            return model.objects.filter(specialityMod__level__departement__est=est).distinct()
        elif model == Schedule:
            return model.objects.filter(speciality__level__departement__est=est).distinct()
        elif model == Department:
            return model.objects.filter(est=est).distinct()
        elif model == AdvancedUserProfile:
            return model.objects.filter(est=est).distinct()
        return model.objects.all()

    if request.user.role == 'dep_admin':
        dep = request.user.depAdminProfile.departement
        if model == Level:
            return model.objects.filter(departement=dep).distinct()
        elif model == Speciality:
            return model.objects.filter(level__departement=dep).distinct()
        elif model == Group:
            return model.objects.filter(speciality__level__departement=dep).distinct()
        elif model == Module:
            return model.objects.filter(specialityMod__level__departement=dep).distinct()
        elif model == Schedule:
            return model.objects.filter(speciality__level__departement=dep).distinct()
        elif model == Department:
            return model.objects.filter(id=dep.id).distinct()
        elif model == AdvancedUserProfile:
            return model.objects.filter(est__department=dep).distinct()
        return model.objects.all()

    if request.user.is_staff and request.user.role == 'advancedUser':
        dep = request.user.Advanceduser_profile.Role.departement
        if model == Level:
            return model.objects.filter(departement=dep).distinct()
        elif model == Speciality:
            return model.objects.filter(level__departement=dep).distinct()
        elif model == Group:
            return model.objects.filter(speciality__level__departement=dep).distinct()
        elif model == Module:
            return model.objects.filter(specialityMod__level__departement=dep).distinct()
        elif model == Schedule:
            return model.objects.filter(speciality__level__departement=dep).distinct()
        elif model == Department:
            return model.objects.filter(id=dep.id).distinct()
        elif model == AdvancedUserProfile:
            return model.objects.filter(est__department=dep).distinct()
        return model.objects.all()

    return model.objects.none()


@admin.register(Level)
class LevelAdmin(admin.ModelAdmin):
    list_display = ('name', 'departement')
    list_filter = ('name', 'departement')
    search_fields = ('name', 'departement__name')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Level)

    def get_form(self, request, obj=None, **kwargs): #personalization of form
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'departement' in form.base_fields:
                form.base_fields['departement'].queryset = get_filtered_queryset(request, Department)
                if request.user.role == 'dep_admin':
                    form.base_fields['departement'].queryset = get_filtered_queryset(request, Department)
                elif request.user.role == 'advancedUser':
                    form.base_fields['departement'].queryset = get_filtered_queryset(request, Department)
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook() #openpyxl: Crée un nouveau fichier Excel (workbook) vide.
            ws = wb.active
            ws.title = "Template"

            headers = ["Département", "Nom du niveau"]
            ws.append(headers)

            departments = get_filtered_queryset(request, Department).values_list('name', flat=True)

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', departments)  # Département

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')#fichier excel avec le format openxml de microsoft à ouvrir avec une application
            response['Content-Disposition'] = 'attachment; filename=Niveaux.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Département': 'departement',
                        'Nom du niveau': 'name',
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            departement_obj = Department.objects.get(name=row['departement']) if row['departement'] else None
                            if not Level.objects.filter(departement=departement_obj, name=row['name']).exists():
                                Level.objects.create(departement=departement_obj, name=row['name'],)
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(Speciality)
class SpecialityAdmin(admin.ModelAdmin):
    list_display = ('name', 'level', 'is_activated')
    list_filter = ('name', 'level', 'level__departement__name', 'is_activated')
    search_fields = ('name', 'level__name', 'level__departement__name')

    actions = ['create_groups', 'assign_students', 'generate_schedule_s1', 'generate_schedule_s2',
               'generate_schedule_no_semester']

    def get_queryset(self, request):
        return get_filtered_queryset(request, Speciality)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'level' in form.base_fields:
                form.base_fields['level'].queryset = get_filtered_queryset(request, Level)
        return form

    def create_groups(self, request, queryset):
        groupCreation(list(queryset))
        self.message_user(request, f"Groupes créés pour {queryset.count()} specialité(s).")

    create_groups.short_description = "Créer des groupes maintenant"

    def assign_students(self, request, queryset):
        assignToGroups(list(queryset))
        self.message_user(request, f"Users affectés pour {queryset.count()} specialité(s).")

    assign_students.short_description = "Affecter les étudiants à des groupes"

    def generate_schedule_s1(self, request, queryset):
        generate_schedule_for_specialities(list(queryset), semester='s1')
        self.message_user(request, f"Planning générée pour S1 dans {queryset.count()} spécialité(s).")

    generate_schedule_s1.short_description = "Générer planning S1 (Universités)"

    def generate_schedule_s2(self, request, queryset):
        generate_schedule_for_specialities(list(queryset), semester='s2')
        self.message_user(request, f"Planning générée pour S2 dans {queryset.count()} spécialité(s).")

    generate_schedule_s2.short_description = "Générer planning S2 (Universités)"

    def generate_schedule_no_semester(self, request, queryset):
        generate_schedule_for_specialities(list(queryset), semester=None)
        self.message_user(request, f"Planning générée pour {queryset.count()} spécialité(s) (Lycées/Écoles).")

    generate_schedule_no_semester.short_description = "Générer planning (Lycées/Écoles)"

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Niveau", "Acronyme de la spécialité", "Nom du specialité", "Nombre de sections"]
            ws.append(headers)

            levels = get_filtered_queryset(request, Level).select_related('departement')
            levels_choices = [
                f"{level.name.strip()} {level.departement.name.strip()[0:3]}".replace(',', '').replace('"', '')
                for level in levels
            ]

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', levels_choices)  # Niveau

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Specialites.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Niveau': 'level',
                        'Acronyme de la spécialité': 'name',
                        'Nom du specialité': 'speciality_nom_complet',
                        "Nombre de sections": "sections",
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None

                            level_str = row['level']
                            if ' ' in level_str:
                                level_name = level_str.split(' ')[0]
                                dep_name = level_str.split(' ')[1]

                            level_obj = Level.objects.annotate(dep_prefix=Substr('departement__name', 1, 3)
                            ).get(name=level_name,dep_prefix=dep_name)
                            if not Speciality.objects.filter(level=level_obj, name=row['name']).exists():
                                Speciality.objects.create(
                                    level=level_obj,
                                    name=row['name'],
                                    speciality_nom_complet=row['speciality_nom_complet'],
                                    sections=row['sections'],
                                    is_activated=True,
                                )
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(Module)
class ModuleAdmin(admin.ModelAdmin):
    list_display = ('module_name', 'module_nom_complet', 'specialityMod', 'coefficient', 'credit',
                    'nbrTDSemaine', 'nbrTPSemaine', 'nbrCoursSemaine', 'semester', 'mode_evaluation', 'is_activated')
    list_filter = ('module_name', 'module_nom_complet', 'specialityMod', 'coefficient', 'credit',
                   'mode_evaluation', 'semester', 'is_activated')
    search_fields = ('module_name', 'module_nom_complet', 'specialityMod__name', 'coefficient',
                     'credit', 'mode_evaluation', 'semester')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Module)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'specialityMod' in form.base_fields:
                form.base_fields['specialityMod'].queryset = get_filtered_queryset(request, Speciality)
            if 'TD' in form.base_fields:
                form.base_fields['TD'].queryset = get_filtered_queryset(request, typeSalle)
            if 'TP' in form.base_fields:
                form.base_fields['TP'].queryset = get_filtered_queryset(request, typeSalle)
            if 'Cours' in form.base_fields:
                form.base_fields['Cours'].queryset = get_filtered_queryset(request, typeSalle)
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = [
                "Spécialité", "Acronyme du module", "Nom du module", "Coefficient", "Crédit",
                "TD ? Type salle", "TP ? Type salle", "Cours ? Type salle", "Nombre TD par Semaine",
                "Nombre TP par Semaine", "Nombre Cours par Semaine", "Nombre d'unités de temps d'examen",
                "Mode d'évaluation", "Semestre"
            ]
            ws.append(headers)

            specialities = get_filtered_queryset(request, Speciality).values_list('name', flat=True)
            room_types = typeSalle.objects.all().values_list('name', flat=True)
            eval_modes = ['Contrôle continu', 'Examen final', 'Contrôle continu + Examen final']
            semesters = ['Semestre 1', 'Semestre 2']

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', specialities)  # Spécialité
            add_dropdown('F', room_types)  # TD Type salle
            add_dropdown('G', room_types)  # TP Type salle
            add_dropdown('H', room_types)  # Cours Type salle
            add_dropdown('M', eval_modes)  # Mode d'évaluation
            add_dropdown('N', semesters)  # Semestre

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Modules.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Spécialité': 'specialityMod',
                        'Acronyme du module': 'module_name',
                        'Nom du module': 'module_nom_complet',
                        'Coefficient': 'coefficient',
                        'Crédit': 'credit',
                        'TD ? Type salle': 'TD',
                        'TP ? Type salle': 'TP',
                        'Cours ? Type salle': 'Cours',
                        'Nombre TD par Semaine': 'nbrTDSemaine',
                        'Nombre TP par Semaine': 'nbrTPSemaine',
                        'Nombre Cours par Semaine': 'nbrCoursSemaine',
                        "Nombre d'unités de temps d'examen": 'examTimeslots',
                        "Mode d'évaluation": 'mode_evaluation',
                        'Semestre': 'semester',
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            speciality_obj = Speciality.objects.get(name=row['specialityMod']) if row['specialityMod'] else None
                            TD_obj = typeSalle.objects.get(name=row['TD']) if row['TD'] else None
                            TP_obj = typeSalle.objects.get(name=row['TP']) if row['TP'] else None
                            Cours_obj = typeSalle.objects.get(name=row['Cours']) if row['Cours'] else None

                            semester_map = {'Semestre 1': 's1', 'Semestre 2': 's2'}
                            semester_value = semester_map.get(row['semester'])
                            if semester_value is None:
                                raise ValueError(f"Semestre invalide: {row['semester']}")

                            modEval_map = {
                                'Contrôle continu': 'Continu',
                                'Examen final': 'Examen',
                                'Contrôle continu + Examen final': 'Both'
                            }
                            modEval_value = modEval_map.get(row['mode_evaluation'])
                            if modEval_value is None:
                                raise ValueError(f"Mode d'evaluation invalide: {row['mode_evaluation']}")

                            if not Module.objects.filter(specialityMod=speciality_obj,
                                                         module_name=row['module_name']).exists():
                                Module.objects.create(
                                    specialityMod=speciality_obj,
                                    module_name=row['module_name'],
                                    module_nom_complet=row['module_nom_complet'],
                                    coefficient=row['coefficient'],
                                    credit=row['credit'],
                                    TD=TD_obj,
                                    TP=TP_obj,
                                    Cours=Cours_obj,
                                    nbrTDSemaine=row['nbrTDSemaine'] if row['nbrTDSemaine'] not in [None,''] else 0,
                                    nbrTPSemaine=row['nbrTPSemaine'] if row['nbrTPSemaine'] not in [None,''] else 0,
                                    nbrCoursSemaine=row['nbrCoursSemaine'] if row['nbrCoursSemaine'] not in [None,''] else 0,
                                    examTimeslots=row['examTimeslots'],
                                    mode_evaluation=modEval_value,
                                    semester=semester_value,
                                    is_activated=True,
                                )
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")


@admin.register(Schedule)
class ScheduleAdmin(admin.ModelAdmin):
    list_display = ('speciality', 'module', 'typeModule', 'section', 'group', 'room_type',
                    'room_number', 'professor', 'day', 'time', 'school_year', 'timestamp', 'is_activated')
    list_filter = ('speciality', 'section', 'group', 'typeModule', 'module__module_name',
                   'room_type', 'room_number', 'professor', 'day', 'school_year', 'timestamp', 'is_activated')
    search_fields = ('module__module_name', 'speciality__name', 'professor__user__first_name',
                     'professor__user__last_name')

    def get_queryset(self, request):
        return get_filtered_queryset(request, Schedule)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'speciality' in form.base_fields:
                form.base_fields['speciality'].queryset = get_filtered_queryset(request, Speciality)
            if 'professor' in form.base_fields:
                form.base_fields['professor'].queryset = get_filtered_queryset(request, AdvancedUserProfile)
            if 'module' in form.base_fields:
                form.base_fields['module'].queryset = get_filtered_queryset(request, Module)
        return form


@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ('speciality', 'group_number', 'capacity')
    list_filter = ('speciality', 'group_number', 'capacity')
    search_fields = ('speciality__name',)

    def get_queryset(self, request):
        return get_filtered_queryset(request, Group)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not request.user.is_superuser:
            if 'speciality' in form.base_fields:
                form.base_fields['speciality'].queryset = get_filtered_queryset(request, Speciality)
        return form

    change_list_template = "Admin/changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def download_template(self, request):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            headers = ["Spécialité", "Numéro du groupe", "Capacité"]
            ws.append(headers)

            specialities = get_filtered_queryset(request, Speciality).values_list('name', flat=True)

            def add_dropdown(column, choices):
                dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"')
                ws.add_data_validation(dv)
                dv.add(f'{column}2:{column}1048576')

            add_dropdown('A', specialities)  # Spécialité

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Groupes.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return HttpResponse("Error while generating Excel", status=500)

    def import_excel(self, request):
        try:
            if request.method == "POST":
                form = ExcelImportForm(request.POST, request.FILES)
                if form.is_valid():
                    verbose_to_field = {
                        'Spécialité': 'speciality',
                        'Numéro du groupe': 'group_number',
                        'Capacité': 'capacity',
                    }

                    try:
                        excel_file = request.FILES['excel_file']
                        #df = pd.read_excel(excel_file)
                        df = pd.read_excel(excel_file, engine='openpyxl')  # Explicit engine

                        df.rename(columns=verbose_to_field, inplace=True)

                        count = 0
                        for _, row in df.iterrows():
                            for each in verbose_to_field.values():
                                if pd.isna(row[each]):
                                    row[each] = None
                            speciality_obj = Speciality.objects.get(name=row['speciality']) if row['speciality'] else None
                            if not Group.objects.filter(speciality=speciality_obj,group_number=row['group_number']).exists():
                                Group.objects.create(speciality=speciality_obj,group_number=row['group_number'],
                                    capacity=row['capacity'],)
                                count += 1
                            else:
                                self.message_user(request, "L'enregistrement existe déjà", level='warning')

                        if count > 0:
                            print(count)
                            self.message_user(request, f"✅ {count} lignes ont été importées avec succès!")
                        return redirect("..")
                    except Exception as e:
                        self.message_user(request, f"Erreur: {str(e)}", level='error')
                        return redirect("..")
            else:
                form = ExcelImportForm()

            return render(request, "Admin/excel_form.html", {
                "form": form,
                "has_download_template": True
            })

        except Exception as e:
            self.message_user(request, "An unexpected error occurred", level='error')
            return redirect("..")